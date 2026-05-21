#!/usr/bin/env python3
"""Create or extend firmware SITL campaign spreadsheet (Intro + phase tabs)."""

from __future__ import print_function

import os
import sys

sys.path.insert(0, os.path.dirname(__file__))
import firmware_SITL_validation_campaign as campaign

P0_ROWS = [
    ('P0-01', 'Parameter load/set', 'Autotest sets and reads core parameters', 'Parameters accepted; no config errors', 'P0_01_Parameters'),
    ('P0-02', 'Arm/disarm and pre-arm', 'Standard arm/disarm sequence and pre-arm checks', 'Clean arm and disarm; pre-arm gates work', 'P0_02_ArmFeatures'),
    ('P0-03', 'Onboard logging', 'Dataflash logging during short flight', 'Log written and readable', 'P0_03_Logging'),
    ('P0-04', 'ALT_HOLD mode', 'Takeoff and hold altitude in ALT_HOLD', 'Stable altitude hold', 'P0_04_ModeAltHold'),
    ('P0-05', 'LOITER mode', 'Takeoff and position hold in LOITER', 'Stable position hold', 'P0_05_ModeLoiter'),
    ('P0-06', 'Takeoff checks', 'Takeoff validation logic', 'Takeoff checks pass', 'P0_06_TakeoffCheck'),
    ('P0-07', 'Landing sequence', 'Standard landing from hover', 'Controlled landing and disarm', 'P0_07_Landing'),
    ('P0-08', 'AUTO mission', 'Fly simple copter AUTO mission', 'Mission completes without fault', 'P0_08_CopterMission'),
    ('P0-09', 'GUIDED sub-modes', 'GUIDED mode sub-mode transitions', 'Predictable guided behaviour', 'P0_09_GuidedSubModeChange'),
    ('P0-10', 'LOITER altitude change', 'Change altitude while in LOITER', 'Reaches target altitude', 'P0_10_LoiterToAlt'),
    ('P0-11', 'Horizontal geofence', 'Circular/polygon horizontal fence', 'Fence breach handled correctly', 'P0_11_HorizontalFence'),
    ('P0-12', 'Throttle failsafe', 'RC throttle failsafe trigger', 'Expected failsafe action', 'P0_12_ThrottleFailsafe'),
    ('P0-13', 'GCS failsafe', 'GCS heartbeat loss failsafe', 'Expected failsafe action', 'P0_13_GCSFailsafe'),
    ('P0-14', 'SMART_RTL', 'SMART_RTL return path', 'Returns home safely', 'P0_14_SMART_RTL'),
    ('P0-15', 'RTL to rally', 'RTL to configured rally point', 'Reaches rally point', 'P0_15_RTL_TO_RALLY'),
    ('P0-16', 'WP nav speed params', 'WPNAV_SPEED parameter behaviour', 'Speed limits respected', 'P0_16_WPNAV_SPEED'),
    ('P0-17', 'Dataflash integrity', 'Log structure and download', 'BIN log valid', 'P0_17_DataFlash'),
    ('P0-18', 'Parameter validation', 'Invalid parameter rejection', 'Bad values rejected', 'P0_18_ParameterChecks'),
    ('P0-19', 'Dijkstra RTL outside inclusion', 'RTL with inclusion fence outside vehicle', 'Path plans and RTL completes', 'P0_19_Dijkstra_OutsideInclusion'),
    ('P0-20', 'Dijkstra RTL inside exclusion', 'RTL from inside exclusion fence', 'Escapes and RTL completes', 'P0_20_Dijkstra_InsideExclusion'),
    ('P0-21', 'Dijkstra path planning return', 'Path planning return to home', 'Valid path and RTL completes', 'P0_21_Dijkstra_PathPlanningReturn'),
    ('P0-22', 'RTL braking distance', 'RTL approach braking behaviour', 'Stops within expected distance', 'P0_22_RTL_BrakingDistance'),
]

SITL_ROWS = [
    ('SITL-01', 'Single polygon exclusion fence', 'AUTO mission path blocked by one polygon exclusion fence', 'Aircraft calculates Dijkstra path around fence and continues mission', '1'),
    ('SITL-02', 'Single circular exclusion fence', 'AUTO mission path blocked by one circular exclusion fence', 'Aircraft routes around circular fence without breach', '1'),
    ('SITL-03', 'Multiple exclusion fences', 'Mission route blocked by two or more exclusion fences', 'Aircraft calculates valid path through available safe corridor', '1'),
    ('SITL-04', 'Inclusion fence only', 'Aircraft operates inside defined inclusion fence', 'Aircraft remains inside allowed area and mission continues normally', '1'),
    ('SITL-05', 'Overlapping exclusion fences', 'Two exclusion fences overlap or touch', 'Aircraft does not generate unsafe path through invalid/blocked area', '1'),
    ('SITL-06', 'Narrow corridor between fences', 'Two fences create a narrow valid route', 'Aircraft only accepts path if clearance is valid against OA_MARGIN_MAX', '1'),
    ('SITL-07', 'No valid path available', 'Fence layout fully blocks route', 'Aircraft fails safely, stops/holds, or triggers expected failsafe behaviour', '1'),
    ('SITL-08', 'Add fence during AUTO mission', 'Start mission with no fence, then upload/enable fence blocking current path', 'Aircraft replans around new fence or safely stops if no path exists', '1'),
    ('SITL-09', 'Change fence during AUTO mission', 'Modify fence geometry while aircraft is flying AUTO', 'Aircraft updates avoidance path without navigation lock-up or unsafe turn', '1'),
    ('SITL-10', 'Delete fence during AUTO mission', 'Remove active blocking fence while Dijkstra avoidance is active', 'Aircraft returns to normal mission path without instability', '1'),
    ('SITL-11', 'Enable fence in air', 'Fence loaded but disabled at takeoff; enable while airborne', 'Fence becomes active and avoidance/geofence logic responds correctly', '1'),
    ('SITL-12', 'Disable fence in air', 'Fence active during mission; disable while airborne', 'Aircraft stops applying fence avoidance and continues normal mission safely', '1'),
    ('SITL-13', 'Fence breach recovery', 'Start inside exclusion fence or create controlled breach', 'Aircraft uses Dijkstra to path-plan out of breached condition and recover', '1'),
    ('SITL-14', 'RTL through blocked path', 'Trigger RTL with exclusion fence between aircraft and home', 'Aircraft avoids fence during RTL path planning', '2'),
    ('SITL-15', 'RTL to rally point through blocked path', 'Rally point is on other side of exclusion fence', 'Aircraft routes around fence while returning to rally point', '1'),
    ('SITL-16', 'Fast waypoints disabled baseline', 'OA_TYPE=2, fast waypoint bit disabled', 'Existing Dijkstra behaviour retained', '1'),
    ('SITL-17', 'Fast waypoints enabled', 'OA_TYPE=2, OA_OPTIONS fast waypoint bit enabled', 'Smoother S-Curve continuation through Dijkstra path points', '2'),
    ('SITL-18', 'Dense generated waypoints', 'Mission has closely spaced waypoints near fence boundary', 'Aircraft does not oscillate, overshoot, or cut into fence', '1'),
    ('SITL-19', 'Waypoints close to fence margin', 'Mission waypoints placed close to exclusion boundary', 'Aircraft maintains required clearance and does not breach fence', '1'),
    ('SITL-20', 'Altitude geofence ceiling', 'Configure max altitude fence and climb mission above limit', 'Aircraft prevents or responds to altitude fence breach as expected', '1'),
    ('SITL-21', 'Altitude + horizontal fence combined', 'Use altitude fence together with polygon/circular exclusion fence', 'Aircraft handles both vertical and horizontal fence constraints safely', '1'),
    ('SITL-22', 'Parameter toggle in flight', 'Toggle OA_OPTIONS fast waypoint bit during airborne mission', 'Behaviour changes predictably or requires restart if expected', '1'),
    ('SITL-23', '3 Close point test', '3 close points with a dogleg at max speed', 'Aircraft follows dogleg through three closely spaced waypoints with fast-waypoint smoothing', '2'),
    ('SITL-25', 'Breach escape vector (RTL failsafe)', 'AUTO/GUIDED flight into circular exclusion; FENCE_ACTION=RTL on breach', 'Aircraft exits via shortest escape vector to safe stand-off; no hover inside fence', '3'),
    ('SITL-26', 'Post-escape RTL home (auto hand-off)', 'After breach escape completes, remain in RTL without mode toggle', 'RTL advances automatically; Dijkstra plans path around fence to home', '3'),
    ('SITL-27', 'No pendulum re-entry after escape', 'RTL after fence breach with fast waypoints enabled', 'Aircraft does not fly back inside exclusion fence after reaching escape stand-off', '3'),
    ('SITL-28', 'Mission WP inside exclusion', 'AUTO mission with waypoint placed inside circular exclusion fence', 'Vehicle holds; path error shown; mission does not silently skip WP', '3'),
    ('SITL-29', 'Guided go-to inside exclusion', 'GUIDED mode; command fly-to point inside exclusion fence', 'Destination rejected; vehicle holds outside exclusion', '2'),
    ('SITL-30', 'Repeated breach RTL cycles', '3+ consecutive breach -> RTL -> recover -> re-mission -> re-breach cycles', 'Each cycle: one escape + one home transit; no stuck hover at stand-off', '3'),
    ('SITL-31', 'Clean GCS messages (release candidate)', 'Breach RTL hand-off with debug diagnostics removed from firmware', 'No debug OADJ messages; only standard fence/Dijkstra messages if applicable', '3'),
    ('SITL-32', 'Mode change at escape stand-off', 'During breach escape, switch RTL -> Loiter -> RTL before home reached', 'Clean recovery; no erratic WP targets or permanent stuck state', '2'),
]

P0_AUTOTEST = {tid: autotest for tid, _, _, _, autotest in P0_ROWS}

FIRMWARE_VERSION = campaign.FIRMWARE_VERSION
REPORT_TITLE = campaign.CAMPAIGN_TITLE

PHASE_SHEET_NAMES = {
    0: campaign.PHASE0_WORKSHEET_NAME,
    1: 'Phase 1 - Fence regression',
    2: 'Phase 2 - Fast-WP integration',
    3: 'Phase 3 - Breach escape gate',
}

PHASE_INFO = [
    {
        'phase': '0',
        'name': 'Firmware update regression',
        'ids': 'P0-01..22',
        'count': '22 tests',
        'scope': (
            'Upstream ArduPilot autotests (parameters, modes, failsafes, RTL, logging) '
            'plus Malloy Dijkstra baseline (outside/inside exclusion, path planning, RTL braking). '
            'Confirms the firmware build is healthy before feature-specific SITL suites run.'
        ),
        'when': 'Every firmware update — run first (generic gate, not tied to a feature).',
        'script': campaign.RUN_TESTS_SCRIPT + ' 0',
        'autotest': 'test.CopterTestsOAfastWPPhase0',
        'worksheet': PHASE_SHEET_NAMES[0],
        'result_cols': 'E–F (Pass/Fail, Log ref on phase tab)',
    },
    {
        'phase': '1',
        'name': 'Dijkstra fence regression',
        'ids': 'SITL-01..23',
        'count': '23 tests',
        'scope': (
            'Full fence and Dijkstra obstacle-avoidance matrix: exclusion/inclusion fences, '
            'dynamic fence changes, breach recovery, RTL/rally, fast-waypoint baseline, '
            'dense waypoints, altitude fences, and in-flight parameter toggles.'
        ),
        'when': 'After Phase 0 passes — broad regression of existing OA behaviour.',
        'script': campaign.RUN_TESTS_SCRIPT + ' 1',
        'autotest': 'test.CopterTestsOAfastWPPhase1',
        'worksheet': PHASE_SHEET_NAMES[1],
        'result_cols': 'E–F (Pass/Fail, Log ref on phase tab)',
    },
    {
        'phase': '2',
        'name': 'Fast-WP integration gate',
        'ids': 'SITL-14, 17, 23, 29, 32',
        'count': '5 tests',
        'scope': (
            'Targeted validation of OA fast waypoints and advanced fence/RTL cases: '
            'RTL through blocked path, fast-waypoint AUTO mission, three-point dogleg smoothing, '
            'GUIDED destination inside exclusion rejected, mode change at breach-escape stand-off.'
        ),
        'when': 'After Phase 1 — confirms fast-WP integration before breach-escape acceptance.',
        'script': campaign.RUN_TESTS_SCRIPT + ' 2',
        'autotest': 'test.CopterTestsOAfastWPPhase2',
        'worksheet': PHASE_SHEET_NAMES[2],
        'result_cols': 'E–F (Pass/Fail, Log ref on phase tab)',
    },
    {
        'phase': '3',
        'name': 'Breach escape gate',
        'ids': 'SITL-25..31 (excl. 29)',
        'count': '6 tests',
        'scope': (
            'Acceptance tests for fenceEscapeVector (4.0.3.16): shortest breach escape vector, '
            'automatic RTL hand-off after escape, no pendulum re-entry, mission WP inside exclusion, '
            'repeated breach/RTL cycles, and clean release GCS messaging (no debug OADJ text).'
        ),
        'when': 'After Phase 2 — sign-off gate for new breach-escape behaviour.',
        'script': campaign.RUN_TESTS_SCRIPT + ' 3',
        'autotest': 'test.CopterTestsOAfastWPPhase3',
        'worksheet': PHASE_SHEET_NAMES[3],
        'result_cols': 'E–F (Pass/Fail, Log ref on phase tab)',
    },
]

PHASE_DATA_START_ROW = 3

SITL_PHASE1_IDS = ['SITL-%02d' % i for i in range(1, 24)]
SITL_PHASE2_IDS = ['SITL-14', 'SITL-17', 'SITL-23', 'SITL-29', 'SITL-32']
SITL_PHASE3_IDS = ['SITL-25', 'SITL-26', 'SITL-27', 'SITL-28', 'SITL-30', 'SITL-31']

SITL_BY_ID = {row[0]: row for row in SITL_ROWS}

PHASE1_AUTOTEST = {
    'SITL-01': 'SITL_01_PolygonExclusionMission',
    'SITL-02': 'SITL_02_CircleExclusionMission',
    'SITL-03': 'SITL_03_MultipleExclusionMission',
    'SITL-04': 'SITL_04_InclusionFenceMission',
    'SITL-05': 'SITL_05_OverlappingExclusionMission',
    'SITL-06': 'SITL_06_NarrowCorridorMission',
    'SITL-07': 'SITL_07_NoValidPathMission',
    'SITL-08': 'SITL_08_AddFenceDuringMission',
    'SITL-09': 'SITL_09_ChangeFenceDuringMission',
    'SITL-10': 'SITL_10_DeleteFenceDuringMission',
    'SITL-11': 'SITL_11_EnableFenceInAir',
    'SITL-12': 'SITL_12_DisableFenceInAir',
    'SITL-13': 'SITL_13_FenceBreachRecovery',
    'SITL-14': 'SITL_14_RTLBlockedPath',
    'SITL-15': 'SITL_15_RTLToRallyBlockedPath',
    'SITL-16': 'SITL_16_FastWaypointsDisabled',
    'SITL-17': 'SITL_17_FastWaypointsEnabled',
    'SITL-18': 'SITL_18_DenseWaypointsNearFence',
    'SITL-19': 'SITL_19_WaypointsCloseToMargin',
    'SITL-20': 'SITL_20_AltitudeGeofenceCeiling',
    'SITL-21': 'SITL_21_AltitudeAndHorizontalFence',
    'SITL-22': 'SITL_22_OAOptionsToggleInFlight',
    'SITL-23': 'SITL_23_ThreePointDogleg',
}

PHASE2_AUTOTEST = {
    'SITL-14': 'OAfastWP_RTL_BlockedPath',
    'SITL-17': 'OAfastWP_FastWaypoints_Mission',
    'SITL-23': 'OAfastWP_ThreePointDogleg',
    'SITL-29': 'OAfastWP_GuidedInsideExclusion',
    'SITL-32': 'OAfastWP_BreachEscape_ModeChangeAtStandoff',
}

PHASE3_AUTOTEST = {
    'SITL-25': 'OAfastWP_BreachEscape_RTL_Home',
    'SITL-26': 'OAfastWP_BreachEscape_RTL_Home',
    'SITL-27': 'OAfastWP_BreachEscape_RTL_Home',
    'SITL-28': 'OAfastWP_MissionWP_InsideExclusion',
    'SITL-30': 'OAfastWP_BreachEscape_RepeatedCycles',
    'SITL-31': 'OAfastWP_BreachEscape_RTL_Home',
}

PHASE_AUTOTEST = {
    0: P0_AUTOTEST,
    1: PHASE1_AUTOTEST,
    2: PHASE2_AUTOTEST,
    3: PHASE3_AUTOTEST,
}

PHASE_TAB_FILLS = {
    0: 'E8F4FC',
    1: 'FFFFFF',
    2: 'FFF4E5',
    3: 'FCE4EC',
}


def _phase_test_ids(phase):
    if phase == 0:
        return [row[0] for row in P0_ROWS]
    if phase == 1:
        return SITL_PHASE1_IDS
    if phase == 2:
        return SITL_PHASE2_IDS
    return SITL_PHASE3_IDS


def _ensure_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        return openpyxl, Alignment, Border, Font, PatternFill, Side
    except ImportError:
        import subprocess
        import sys
        subprocess.check_call(
            [sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q', '--target', '/tmp/pylibs'])
        sys.path.insert(0, '/tmp/pylibs')
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        return openpyxl, Alignment, Border, Font, PatternFill, Side


def _write_intro_sheet(wb, active_phases):
    openpyxl, Alignment, Border, Font, PatternFill, Side = _ensure_openpyxl()
    if 'Intro' in wb.sheetnames:
        del wb['Intro']
    ws = wb.create_sheet('Intro', 0)

    active_phases = [int(p) for p in active_phases]
    active_info = [PHASE_INFO[p] for p in active_phases]

    title_font = Font(bold=True, size=18, color='1F3864')
    subtitle_font = Font(bold=True, size=12, color='4472C4')
    section_font = Font(bold=True, size=11, color='1F3864')
    body_font = Font(size=11)
    table_header_fill = PatternFill('solid', fgColor='4472C4')
    table_header_font = Font(bold=True, color='FFFFFF', size=10)
    phase_fills = {
        '0': PatternFill('solid', fgColor='E8F4FC'),
        '1': PatternFill('solid', fgColor='FFFFFF'),
        '2': PatternFill('solid', fgColor='FFF4E5'),
        '3': PatternFill('solid', fgColor='FCE4EC'),
    }
    wrap = Alignment(wrap_text=True, vertical='top')
    thin = Side(style='thin', color='B4C6E7')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:I1')
    ws['A1'] = REPORT_TITLE
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(vertical='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = (
        'THISFIRMWARE: %s  |  Branch: %s  |  docs/%s/'
        % (campaign.THISFIRMWARE or campaign.CAMPAIGN_ID, campaign.GIT_BRANCH, campaign.CAMPAIGN_ID)
    )
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = wrap

    ws.merge_cells('A4:I4')
    ws['A4'] = (
        'This workbook records unattended ArduCopter SITL autotest results for firmware %s. '
        'All evidence for this campaign lives under docs/%s/ (logs, visual evidence, reports). '
        'Phase 0 is the generic new-firmware regression gate (P0-01..22) — run on every update. '
        'Additional phase worksheets are added when validating feature-specific behaviour.'
        % (FIRMWARE_VERSION, campaign.CAMPAIGN_ID)
    )
    ws['A4'].font = body_font
    ws['A4'].alignment = wrap

    ws.merge_cells('A6:I6')
    ws['A6'] = 'Test phases (cover summary)'
    ws['A6'].font = section_font

    phase_headers = [
        'Phase', 'Name', 'Worksheet tab', 'Test IDs', 'Count', 'Scope / purpose', 'When to run',
        'Run script', 'Autotest class',
    ]
    header_row = 7
    for c, label in enumerate(phase_headers, 1):
        cell = ws.cell(header_row, c, label)
        cell.fill = table_header_fill
        cell.font = table_header_font
        cell.alignment = wrap
        cell.border = border

    row = header_row + 1
    for info in active_info:
        values = [
            info['phase'], info['name'], info['worksheet'], info['ids'], info['count'], info['scope'],
            info['when'], info['script'], info['autotest'],
        ]
        fill = phase_fills.get(info['phase'])
        for c, value in enumerate(values, 1):
            cell = ws.cell(row, c, value)
            cell.font = body_font
            cell.alignment = wrap
            cell.border = border
            if fill is not None:
                cell.fill = fill
        row += 1

    pending = [p for p in range(4) if p not in active_phases]
    if pending:
        row += 1
        ws.merge_cells('A%d:I%d' % (row, row))
        ws.cell(row, 1, 'Planned phases (not yet in workbook)').font = section_font
        row += 1
        ws.merge_cells('A%d:I%d' % (row, row))
        lines = []
        for p in pending:
            info = PHASE_INFO[p]
            lines.append('Phase %s — %s (%s)' % (info['phase'], info['name'], info['ids']))
        lines.append('Add with: %s <phase>' % campaign.ADD_PHASE_SCRIPT)
        ws.cell(row, 1, '\n'.join(lines)).font = body_font
        ws.cell(row, 1).alignment = wrap

    row += 2
    ws.merge_cells('A%d:I%d' % (row, row))
    ws.cell(row, 1, 'Campaign workflow').font = section_font
    row += 1
    ws.merge_cells('A%d:I%d' % (row, row))
    ws.cell(row, 1, (
        '0. New firmware:         update THISFIRMWARE in ArduCopter/version.h, then %s\n'
        '1. Run autotests:        %s <phase>\n'
        '2. Generate artifacts:   %s <phase>\n'
        'Add later phases:        %s <1|2|3>'
        % (campaign.RESET_SCRIPT, campaign.RUN_TESTS_SCRIPT,
           campaign.GENERATE_ARTIFACTS_SCRIPT, campaign.ADD_PHASE_SCRIPT)
    )).font = body_font
    ws.cell(row, 1).alignment = wrap

    row += 2
    ws.merge_cells('A%d:I%d' % (row, row))
    ws.cell(row, 1, 'Evidence layout (under docs/%s/)' % campaign.CAMPAIGN_ID).font = section_font
    row += 1
    ws.merge_cells('A%d:I%d' % (row, row))
    ws.cell(row, 1, (
        'phase<N>/logs/              autotest .txt / .tlog / .BIN output\n'
        'phase<N>/visual_evidence/   HTML dashboard + PNG result cards\n'
        'phase0/report/              firmware regression RTF + HTML (%s)\n'
        'phase<N>/report/            phase HTML report (phases 1–3)\n'
        'Spreadsheet columns E–F: Pass/Fail and log ref on each phase tab.'
        % campaign.FIRMWARE_REGRESSION_RTF_NAME
    )).font = body_font
    ws.cell(row, 1).alignment = wrap

    widths = {'A': 8, 'B': 24, 'C': 22, 'D': 16, 'E': 10, 'F': 40, 'G': 24, 'H': 30, 'I': 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[4].height = 48
    for r in range(header_row + 1, header_row + 1 + len(active_info)):
        ws.row_dimensions[r].height = 72
    ws.sheet_view.showGridLines = False
    return ws


def _write_phase_sheet(wb, phase):
    openpyxl, Alignment, _, Font, PatternFill, _ = _ensure_openpyxl()
    info = PHASE_INFO[phase]
    sheet_name = PHASE_SHEET_NAMES[phase]
    ws = wb.create_sheet(sheet_name)

    title_font = Font(bold=True, size=12, color='1F3864')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF')
    tab_fill = PatternFill('solid', fgColor=PHASE_TAB_FILLS[phase])
    wrap = Alignment(wrap_text=True, vertical='top')

    ws.merge_cells('A1:H1')
    ws['A1'] = 'Phase %s — %s (%s)' % (info['phase'], info['name'], info['ids'])
    ws['A1'].font = title_font
    ws['A1'].alignment = wrap
    ws['A1'].fill = tab_fill

    headers = [
        'Test ID', 'Test Case', 'Setup / Action', 'Expected Result',
        'Pass / Fail', 'Log ref', 'Firmware', 'Autotest method',
    ]
    header_row = 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(header_row, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    autotest_map = PHASE_AUTOTEST[phase]
    row = PHASE_DATA_START_ROW
    if phase == 0:
        for tid, case, setup, expected, autotest in P0_ROWS:
            ws.cell(row, 1, tid)
            ws.cell(row, 2, case)
            ws.cell(row, 3, setup)
            ws.cell(row, 4, expected)
            ws.cell(row, 7, FIRMWARE_VERSION)
            ws.cell(row, 8, autotest)
            for c in range(1, 9):
                ws.cell(row, c).alignment = wrap
                if phase == 0:
                    ws.cell(row, c).fill = tab_fill
            row += 1
    else:
        for tid in _phase_test_ids(phase):
            _, case, setup, expected, _ = SITL_BY_ID[tid]
            ws.cell(row, 1, tid)
            ws.cell(row, 2, case)
            ws.cell(row, 3, setup)
            ws.cell(row, 4, expected)
            ws.cell(row, 7, FIRMWARE_VERSION)
            ws.cell(row, 8, autotest_map.get(tid, ''))
            for c in range(1, 9):
                ws.cell(row, c).alignment = wrap
            row += 1

    widths = {'A': 10, 'B': 28, 'C': 42, 'D': 42, 'E': 14, 'F': 36, 'G': 10, 'H': 32}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A%d' % PHASE_DATA_START_ROW
    return ws


def _active_phases_in_workbook(wb):
    phases = []
    for p, name in PHASE_SHEET_NAMES.items():
        if name in wb.sheetnames:
            phases.append(p)
    return sorted(phases)


def write_spreadsheet(path, phases=(0,)):
    openpyxl, _, _, _, _, _ = _ensure_openpyxl()
    phases = sorted(set(int(p) for p in phases))

    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    for phase in phases:
        _write_phase_sheet(wb, phase)
    _write_intro_sheet(wb, phases)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


def add_phases_to_spreadsheet(path, phases):
    openpyxl, _, _, _, _, _ = _ensure_openpyxl()
    phases = sorted(set(int(p) for p in phases))

    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path)
    else:
        return write_spreadsheet(path, phases)

    existing = set(_active_phases_in_workbook(wb))
    for phase in phases:
        if phase not in existing:
            _write_phase_sheet(wb, phase)
    all_phases = sorted(existing | set(phases))
    _write_intro_sheet(wb, all_phases)
    wb.save(path)
    return path


def init_campaign_spreadsheet():
    ensure_campaign_dirs(phases=(0,))
    path = campaign.spreadsheet_path()
    write_spreadsheet(path, phases=(0,))
    return path


def ensure_campaign_dirs(phases=(0,)):
    campaign.ensure_campaign_dirs(phases=phases)


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == 'add':
        phase = int(sys.argv[2])
        path = add_phases_to_spreadsheet(campaign.spreadsheet_path(), (phase,))
        print('Extended spreadsheet:', path)
    else:
        path = init_campaign_spreadsheet()
        print('Wrote', path)
