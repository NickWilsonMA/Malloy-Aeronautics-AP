#!/usr/bin/env python3
"""Fill campaign spreadsheet from autotest logs under docs/<campaign>/Phase<N>/logs/."""

from __future__ import print_function

import argparse
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q', '--target', '/tmp/pylibs'])
    sys.path.insert(0, '/tmp/pylibs')
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.dirname(__file__))
import generate_oafastwp_report_common as rpt
import firmware_SITL_validation_campaign as campaign
import oafastwp_spreadsheet_data as sheet

PASS_COL = 'E'
RERUN_COL = 'F'
NOTE_COL = 'G'
PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_ROW = 2


def result_for_test(buildlogs, test_name, phase):
    txt = rpt.latest_test_log_for_campaign(buildlogs, test_name)
    if not txt:
        return None, None, None
    passed, content = rpt.parse_test_result(txt)
    note = campaign.log_ref_relative_path(txt)
    m = re.search(r'FAILED: "[^"]+": ([^(]+)', content)
    if not passed and m:
        note += '; %s' % m.group(1).strip()
    reruns, got_pass = campaign.reruns_to_pass(buildlogs, test_name)
    if got_pass:
        rerun_display = reruns
    else:
        rerun_display = '—'
    return ('Pass' if passed else 'Fail'), rerun_display, note


def ensure_rerun_column(ws):
    '''Insert Re-runs column F on legacy sheets (Pass/Fail | Log ref | ...).'''
    header = ws.cell(HEADER_ROW, 6).value
    if header and str(header).strip() == 'Re-runs':
        return
    if header and str(header).strip() == 'Log ref':
        ws.insert_cols(6)
        cell = ws.cell(HEADER_ROW, 6, 'Re-runs')
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 36


def update_phase(ws, buildlogs, phase, mapping):
    ensure_rerun_column(ws)
    updated = 0
    for r in range(sheet.PHASE_DATA_START_ROW, ws.max_row + 1):
        tid = ws.cell(r, 1).value
        if not tid:
            continue
        tid = str(tid).strip()
        if tid not in mapping:
            continue
        status, reruns, note = result_for_test(buildlogs, mapping[tid], phase)
        if status is None:
            continue
        ws['%s%d' % (PASS_COL, r)] = status
        ws['%s%d' % (RERUN_COL, r)] = reruns
        ws['%s%d' % (NOTE_COL, r)] = note
        ws['%s%d' % (PASS_COL, r)].fill = PASS_FILL if status == 'Pass' else FAIL_FILL
        ws['%s%d' % (PASS_COL, r)].alignment = Alignment(horizontal='center', vertical='top')
        ws['%s%d' % (RERUN_COL, r)].alignment = Alignment(horizontal='center', vertical='top')
        ws['%s%d' % (NOTE_COL, r)].alignment = Alignment(wrap_text=True, vertical='top')
        updated += 1
    return updated


def update_all_spreadsheet(xlsx=None, phases=None):
    '''Update every phase worksheet that exists in the workbook and has logs on disk.'''
    xlsx = xlsx or campaign.spreadsheet_path()
    if not os.path.isfile(xlsx):
        print('Spreadsheet not found: %s' % xlsx, file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx)
    workbook_phases = sheet._active_phases_in_workbook(wb)
    if phases is None:
        phases = workbook_phases
    else:
        phases = [int(p) for p in phases]

    total = 0
    for phase in phases:
        sheet_name = sheet.PHASE_SHEET_NAMES[phase]
        if sheet_name not in wb.sheetnames:
            print('Phase %d: skipped (worksheet %s missing)' % (phase, sheet_name))
            continue
        buildlogs = os.path.abspath(campaign.phase_buildlogs(phase))
        if not os.path.isdir(buildlogs):
            print('Phase %d: skipped (no logs at %s)' % (phase, buildlogs))
            continue
        ws = wb[sheet_name]
        n = update_phase(ws, buildlogs, phase, sheet.PHASE_AUTOTEST[phase])
        print('Phase %d (%s): updated %d rows from %s' % (phase, sheet_name, n, buildlogs))
        total += n

    wb.save(xlsx)
    print('Spreadsheet saved: %s (%d rows updated across %d phase(s))' % (xlsx, total, len(phases)))
    return 0


def update_spreadsheet(phase, buildlogs, xlsx=None):
    xlsx = xlsx or campaign.spreadsheet_path()
    if not os.path.isfile(xlsx):
        print('Spreadsheet not found: %s' % xlsx, file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx)
    sheet_name = sheet.PHASE_SHEET_NAMES[phase]
    if sheet_name not in wb.sheetnames:
        print('Phase %d: worksheet missing (%s)' % (phase, sheet_name), file=sys.stderr)
        return 1

    buildlogs = os.path.abspath(buildlogs)
    ws = wb[sheet_name]
    n = update_phase(ws, buildlogs, phase, sheet.PHASE_AUTOTEST[phase])
    wb.save(xlsx)
    print('Spreadsheet: updated %d rows -> %s' % (n, xlsx))
    return 0


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--xlsx', default=campaign.spreadsheet_path())
    parser.add_argument('--buildlogs', default=None)
    parser.add_argument('--phase', type=int, choices=[0, 1, 2, 3], default=None,
                        help='Phase to update (omit for all sheets present in workbook)')
    parser.add_argument('--rebuild', action='store_true',
                        help='Rebuild spreadsheet structure for phases already in workbook')
    args = parser.parse_args()

    if args.rebuild:
        if not os.path.isfile(args.xlsx):
            sheet.init_campaign_spreadsheet()
        else:
            wb = openpyxl.load_workbook(args.xlsx)
            phases = sheet._active_phases_in_workbook(wb)
            sheet.write_spreadsheet(args.xlsx, phases=phases or (0,))
        print('Rebuilt spreadsheet structure: %s' % args.xlsx)

    if not os.path.isfile(args.xlsx):
        print('Spreadsheet not found. Run reset_firmware_SITL_validation_campaign.sh first.', file=sys.stderr)
        return 1

    if args.phase is not None:
        buildlogs = os.path.abspath(args.buildlogs or campaign.phase_buildlogs(args.phase))
        if not os.path.isdir(buildlogs):
            print('No logs at %s' % buildlogs, file=sys.stderr)
            return 1
        return update_spreadsheet(args.phase, buildlogs, args.xlsx)

    return update_all_spreadsheet(args.xlsx)


if __name__ == '__main__':
    sys.exit(main())
